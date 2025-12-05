from rest_framework import viewsets, filters
from .models import User
from .models_db import *
from .serializer import *
from .forms import *
from rest_framework import generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth import authenticate
from rest_framework.permissions import IsAuthenticated, IsAuthenticatedOrReadOnly
from rest_framework.decorators import action
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from io import BytesIO
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
import datetime
try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

from rest_framework_simplejwt.views import TokenObtainPairView
from .serializer import CustomTokenObtainPairSerializer
import django_filters


class AcudienteUserMatchViewSet(viewsets.ViewSet):
    """
    Busca coincidencia entre el username del usuario (rol=padres)
    y el número de documento del acudiente.
    Si encuentra el acudiente, trae los estudiantes relacionados.
    """

    @action(detail=False, methods=['get'])
    def verificar_coincidencias(self, request):
        # Obtener el número de documento desde el query param
        numero_doc = request.query_params.get('numero_documento_acudiente', '').strip()

        if not numero_doc:
            return Response({
                "error": "Debe enviar el parámetro 'numero_documento_acudiente'. Ejemplo: ?numero_documento_acudiente=123456789"
            }, status=status.HTTP_400_BAD_REQUEST)

        # Buscar acudiente y usuario por coincidencia de documento/username
        acudiente = Acudiente.objects.filter(numero_documento_acudiente=numero_doc).first()
        usuario = User.objects.filter(username=numero_doc, rol='padres').first()

        # Buscar estudiantes relacionados si hay acudiente
        estudiantes_data = []
        if acudiente:
            relaciones = EstudiantesAcudientes.objects.filter(
                fk_numero_documento_acudiente=acudiente
            ).select_related('fk_numero_documento_estudiante')

            for rel in relaciones:
                est = rel.fk_numero_documento_estudiante
                estudiantes_data.append({
                    "numero_documento": est.numero_documento_estudiante,
                    "nombre_completo": est.nombre_completo,
                    "correo": est.correo,
                    "edad": est.edad,
                    "telefono": est.telefono,
                    "direccion": est.direccion,
                })

        # Construcción de respuesta
        respuesta = {}
        if acudiente and usuario:
            respuesta.update({
                "mensaje": "Coincidencia encontrada.",
                "acudiente": acudiente.nombre_completo,
                "numero_documento_acudiente": acudiente.numero_documento_acudiente,
                "usuario": f"{usuario.first_name} {usuario.last_name}".strip(),
                "username": usuario.username,
                "rol": usuario.rol
            })
        elif acudiente:
            respuesta.update({
                "mensaje": "Solo se encontró coincidencia en acudiente.",
                "acudiente": acudiente.nombre_completo,
                "numero_documento_acudiente": acudiente.numero_documento_acudiente
            })
        elif usuario:
            respuesta.update({
                "mensaje": "Solo se encontró coincidencia en usuario (rol padres).",
                "usuario": f"{usuario.first_name} {usuario.last_name}".strip(),
                "username": usuario.username
            })
        else:
            return Response({"mensaje": "No se encontró coincidencia con el número de documento proporcionado."})

        respuesta["estudiantes_relacionados"] = estudiantes_data or []

        return Response(respuesta)

class TipoDocumentoViewSet(viewsets.ModelViewSet):
    queryset = TipoDocumento.objects.all()
    serializer_class = TipoDocumentoSerializer

class anoElectivoViewSet(viewsets.ModelViewSet):
    queryset = ano_electivo.objects.all()
    serializer_class = AnoElectivoSerializer

class TipoActividadViewSet(viewsets.ModelViewSet):
    queryset = TipoActividad.objects.all()
    serializer_class = TipoActividadSerializer

class ActividadesFilter(django_filters.FilterSet):
    profesor = django_filters.CharFilter(
        field_name='fk_id_materia_profesores__fk_numero_documento_profesor__numero_documento_profesor',
        lookup_expr='exact'
    )

    class Meta:
        model = Actividades
        fields = ['profesor']

class ActividadesViewSet(viewsets.ModelViewSet):
    queryset = Actividades.objects.all()
    serializer_class = ActividadesSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = ActividadesFilter

class TipoSangreViewSet(viewsets.ModelViewSet):
    queryset = TipoSangre.objects.all()
    serializer_class = TipoSangreSerializer


class GeneroViewSet(viewsets.ModelViewSet):
    queryset = Genero.objects.all()
    serializer_class = GeneroSerializer


class EstadoViewSet(viewsets.ModelViewSet):
    queryset = TipoEstado.objects.all()
    serializer_class = TipoEstadoSerializer


class DepartamentoViewSet(viewsets.ModelViewSet):
    queryset = Departamento.objects.all()
    serializer_class = DepartamentoSerializer


class CiudadViewSet(viewsets.ModelViewSet):
    queryset = Ciudad.objects.all()
    serializer_class = CiudadSerializer

class MateriasAsignadasViewSet(viewsets.ModelViewSet):
    queryset = MateriasAsignadas.objects.all()
    serializer_class = MateriasAsignadasSerializer

    filter_backends = [DjangoFilterBackend, filters.SearchFilter]

    # Exacto: solo devuelve lo que coincida 100%
    filterset_fields = ['fk_numero_documento_profesor']

    # Parcial: devuelve coincidencias parciales
    search_fields = [
        'fk_id_materia__nombre', 
    ]

class EstudiantesCursosViewSet(viewsets.ModelViewSet):
    queryset = Estudiantes_cursos.objects.all()
    serializer_class = EstudiantesCursosSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['numero_documento_estudiante__numero_documento_estudiante', 'id_curso__nombre']

class AreaConocimientoViewSet(viewsets.ModelViewSet):
    queryset = Area_conocimiento.objects.all()
    serializer_class = AreaConocimientoSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre']

class MateriasViewSet(viewsets.ModelViewSet):
    queryset = Materias.objects.all()
    serializer_class = MatriaSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre']

class EstudianteViewSet(viewsets.ModelViewSet):
    queryset = Estudiantes.objects.all()
    serializer_class = EstudiantesSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['numero_documento_estudiante']

class CursoViewSet(viewsets.ModelViewSet):
    queryset = Cursos.objects.all()
    serializer_class = CursoSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre']

class SisbenViewSet(viewsets.ModelViewSet):
    queryset = Sisben.objects.all()
    serializer_class = SisbenSerializer

class DiscapacidadViewSet(viewsets.ModelViewSet):
    queryset = Discapacidad.objects.all()
    serializer_class = DiscapacidadSerializer

class AlergiaViewSet(viewsets.ModelViewSet):
    queryset = Alergia.objects.all()
    serializer_class = AlergiaSerializer

class ProfesorViewSet(viewsets.ModelViewSet):
    queryset = Profesores.objects.all()
    serializer_class = ProfesoresSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['numero_documento_profesor']


class AcudienteViewSet(viewsets.ModelViewSet):
    queryset = Acudiente.objects.all()
    serializer_class = AcudienteSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['numero_documento_acudiente']

class EstudianteNotasViewSet(viewsets.ModelViewSet):
    queryset = EstudianteNotas.objects.all()
    serializer_class = EstudianteNotasSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['fk_numero_documento_estudiante']


class EstudianteAcudienteViewSet(viewsets.ModelViewSet):
    queryset = EstudiantesAcudientes.objects.all()
    serializer_class = EstudiantesAcudientesSerializer

class EstudiantesTemplateView(APIView):
    permission_classes = [IsAuthenticatedOrReadOnly]

    def get(self, request):
        print("🔥 DEBUG: EstudiantesTemplateView ejecutándose - NUEVA VERSION")
        if Workbook is None:
            return Response({
                'error': 'openpyxl no está instalado. Instala con: pip install openpyxl'
            }, status=500)

        wb = Workbook()
        # Hoja Estudiantes (solo cabeceras, sin datos de ejemplo)
        ws_est = wb.active
        ws_est.title = 'Estudiantes'
        headers_est = [
            'numero_documento_estudiante', 'tipo_documento', 'nombre1', 'nombre2',
            'apellido1', 'apellido2', 'correo', 'direccion', 'genero',
            'fecha_nacimiento(YYYY-MM-DD)', 'municipio', 'telefono',
            'tipo_sangre', 'sisben', 'estado', 'religion'
        ]
        ws_est.append(headers_est)

        # Hoja Acudientes (solo cabeceras, sin datos de ejemplo)
        ws_acu = wb.create_sheet(title='Acudientes')
        headers_acu = [
            'numero_documento_acudiente', 'tipo_documento', 'nombre1', 'nombre2',
            'apellido1', 'apellido2', 'telefono1', 'telefono2', 'direccion',
            'fk_codigo_municipio', 'numero_documento_estudiante', 'tipo_acudiente'
        ]
        ws_acu.append(headers_acu)

        # Hoja CATALOGO solo con Ciudad ID/DESCRIPCION como muestra la imagen
        ws_cat = wb.create_sheet(title='CATALOGO')
        ws_cat.append(['ID', 'DESCRIPCION'])
        
        # Obtener municipios y llenar la hoja CATALOGO
        municipios_qs = Ciudad.objects.select_related('fk_codigo_departamento').all()
        for c in municipios_qs:
            ws_cat.append([c.codigo_municipio, c.nombre])

        # Hoja "Instrucciones" con layout de tablas como en la imagen
        ws_inst = wb.create_sheet(title='Instrucciones')
        
        # Tabla id_tipo_sisben (columna A-B)
        ws_inst['A1'] = 'id_tipo_sisben'
        ws_inst['B1'] = 'descripcion'
        sisben_data = Sisben.objects.all()
        for i, ss in enumerate(sisben_data, start=2):
            ws_inst[f'A{i}'] = getattr(ss, 'id_tipo_sisben', getattr(ss, 'pk', None))
            ws_inst[f'B{i}'] = ss.descripcion

        # Tabla id_tipo_sangre (columna D-E)
        ws_inst['D1'] = 'id_tipo_sangre'
        ws_inst['E1'] = 'descripcion'
        sangre_data = TipoSangre.objects.all()
        for i, s in enumerate(sangre_data, start=2):
            ws_inst[f'D{i}'] = getattr(s, 'id_tipo_sangre', getattr(s, 'pk', None))
            ws_inst[f'E{i}'] = s.descripcion

        # Tabla id_tipo_acudiente (columna G-H)
        ws_inst['G1'] = 'id_tipo_acudiente'
        ws_inst['H1'] = 'descripcion'
        acudiente_data = TipoAcudiente.objects.all()
        for i, ta in enumerate(acudiente_data, start=2):
            ws_inst[f'G{i}'] = getattr(ta, 'id_tipo_acudiente', getattr(ta, 'pk', None))
            ws_inst[f'H{i}'] = ta.descripcion

        # Tabla id_tipo_discapacidad (columna J-K)
        ws_inst['J1'] = 'id_tipo_discapacidad'
        ws_inst['K1'] = 'descripcion'
        # Asumiendo que tienes modelo Discapacidad, si no ajustar
        try:
            from .models_db import Discapacidad
            discapacidad_data = Discapacidad.objects.all()
            for i, d in enumerate(discapacidad_data, start=2):
                ws_inst[f'J{i}'] = getattr(d, 'id_tipo_discapacidad', getattr(d, 'pk', None))
                ws_inst[f'K{i}'] = d.descripcion
        except:
            # Si no existe el modelo, llenar con datos de ejemplo
            ws_inst['J2'] = 1
            ws_inst['K2'] = 'Ninguna'
            ws_inst['J3'] = 2
            ws_inst['K3'] = 'Motriz'
            ws_inst['J4'] = 3
            ws_inst['K4'] = 'Visual'
            ws_inst['J5'] = 4
            ws_inst['K5'] = 'Auditiva'
            ws_inst['J6'] = 5
            ws_inst['K6'] = 'Intelectual'
            ws_inst['J7'] = 6
            ws_inst['K7'] = 'Psicosocial'
            ws_inst['J8'] = 7
            ws_inst['K8'] = 'Comunicativa'
            ws_inst['J9'] = 8
            ws_inst['K9'] = 'Múltiple'

        # Segunda fila de tablas
        # Buscar la fila donde termina la tabla más larga de arriba
        max_row_first = max(
            len(sisben_data) + 1,
            len(sangre_data) + 1,
            len(acudiente_data) + 1,
            9  # Para discapacidad
        ) + 3  # Espacio

        # Tabla id_tipo_alergia (columna A-B)
        ws_inst[f'A{max_row_first}'] = 'id_tipo_alergia'
        ws_inst[f'B{max_row_first}'] = 'descripcion'
        # Datos de ejemplo para alergia
        alergia_ejemplos = [
            (1, 'Ninguna'), (2, 'Polen'), (3, 'Maní'), (4, 'Lácteos'),
            (5, 'Mariscos'), (6, 'Huevo'), (7, 'Soya'), (8, 'Gluten'),
            (9, 'Medicamentos'), (10, 'Insectos')
        ]
        for i, (id_val, desc) in enumerate(alergia_ejemplos, start=max_row_first + 1):
            ws_inst[f'A{i}'] = id_val
            ws_inst[f'B{i}'] = desc

        # Tabla id_tipo_rol (columna D-E)
        ws_inst[f'D{max_row_first}'] = 'id_tipo_rol'
        ws_inst[f'E{max_row_first}'] = 'descripcion'
        rol_ejemplos = [
            (1, 'Administrador'), (2, 'Docente'), (3, 'Estudiante'), (4, 'Acudiente')
        ]
        for i, (id_val, desc) in enumerate(rol_ejemplos, start=max_row_first + 1):
            ws_inst[f'D{i}'] = id_val
            ws_inst[f'E{i}'] = desc

        # Tabla id_tipo_documento (columna G-H)
        ws_inst[f'G{max_row_first}'] = 'id_tipo_documento'
        ws_inst[f'H{max_row_first}'] = 'descripcion'
        documento_data = TipoDocumento.objects.all()
        for i, t in enumerate(documento_data, start=max_row_first + 1):
            ws_inst[f'G{i}'] = getattr(t, 'id_tipo_documento', getattr(t, 'pk', None))
            ws_inst[f'H{i}'] = t.descripcion

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_estudiantes.xlsx"'
        return response
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_estudiantes.xlsx"'
        return response

class ProfesoresTemplateView(APIView):
    permission_classes = [IsAuthenticatedOrReadOnly]

    def get(self, request):
        if Workbook is None:
            return Response({
                'error': 'openpyxl no está instalado. Instala con: pip install openpyxl'
            }, status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Profesores'
        headers = [
            'numero_documento_profesor', 'tipo_documento', 'nombre1', 'nombre2',
            'apellido1', 'apellido2', 'correo', 'direccion', 'telefono1', 'telefono2',
            'estado', 'municipio', 'tipo_sangre'
        ]
        ws.append(headers)
        ws.append([
            '2000001', 'Cédula de ciudadanía', 'María', '', 'López', '', 'maria@example.com', 'Carrera 45 #10-20',
            '3100000000', '', 'Activo', 'Bogotá', 'A+'
        ])

        # Hoja de Catálogos con valores permitidos
        ws_cat = wb.create_sheet(title='Catalogos')
        ws_cat.append(['TIPO_DOCUMENTO', 'ESTADO', 'TIPO_SANGRE', 'MUNICIPIO'])
        
        # Obtener valores de catálogos
        tipos_doc = list(TipoDocumento.objects.values_list('descripcion', flat=True))
        estados = list(TipoEstado.objects.values_list('descripcion', flat=True))
        tipos_sangre = list(TipoSangre.objects.values_list('descripcion', flat=True))
        municipios_qs = Ciudad.objects.select_related('fk_codigo_departamento').all()
        municipios = [f"{c.codigo_municipio} - {c.nombre}" for c in municipios_qs]

        # Llenar filas con valores de catálogos
        max_len = max(len(tipos_doc), len(estados), len(tipos_sangre), len(municipios))
        for i in range(max_len):
            row = [
                tipos_doc[i] if i < len(tipos_doc) else '',
                estados[i] if i < len(estados) else '',
                tipos_sangre[i] if i < len(tipos_sangre) else '',
                municipios[i] if i < len(municipios) else ''
            ]
            ws_cat.append(row)

        # Hoja con detalle por ID para cada catálogo
        ws_det = wb.create_sheet(title='Catalogos_Detalle')
        ws_det.append(['TIPO_DOCUMENTO_ID', 'TIPO_DOCUMENTO_DESCRIPCION'])
        for t in TipoDocumento.objects.all():
            ws_det.append([getattr(t, 'id_tipo_documento', getattr(t, 'pk', None)), t.descripcion])
        ws_det.append([])

        ws_det.append(['ID_TIPO_ESTADO', 'TIPO_ESTADO_DESCRIPCION'])
        for e in TipoEstado.objects.all():
            ws_det.append([getattr(e, 'id_tipo_estado', getattr(e, 'pk', None)), e.descripcion])
        ws_det.append([])

        ws_det.append(['ID_TIPO_SANGRE', 'TIPO_SANGRE_DESCRIPCION'])
        for s in TipoSangre.objects.all():
            ws_det.append([getattr(s, 'id_tipo_sangre', getattr(s, 'pk', None)), s.descripcion])
        ws_det.append([])

        ws_det.append(['CODIGO_MUNICIPIO', 'NOMBRE_MUNICIPIO', 'CODIGO_DEPARTAMENTO', 'NOMBRE_DEPARTAMENTO'])
        for c in municipios_qs:
            depto_code = getattr(c.fk_codigo_departamento, 'codigo_departamento', None) if c.fk_codigo_departamento else None
            depto_name = getattr(c.fk_codigo_departamento, 'nombre', '') if c.fk_codigo_departamento else ''
            ws_det.append([c.codigo_municipio, c.nombre, depto_code, depto_name])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_profesores.xlsx"'
        return response


class EstudiantesBulkUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if load_workbook is None:
            return Response({'error': 'openpyxl no está instalado. pip install openpyxl'}, status=500)
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Archivo "file" es requerido'}, status=400)

        try:
            wb = load_workbook(filename=file, data_only=True)
            ws = wb['Estudiantes'] if 'Estudiantes' in wb.sheetnames else wb.active
        except Exception as e:
            return Response({'error': f'Error leyendo Excel: {str(e)}'}, status=400)

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        # Aceptar nombres naturales además de IDs para columnas de FK
        required = ['numero_documento_estudiante', 'nombre1', 'apellido1']
        for req in required:
            if req not in headers:
                return Response({'error': f'Columna requerida faltante: {req}'}, status=400)
        # tipo_documento puede venir como fk_id_tipo_documento o tipo_documento (nombre)
        if ('fk_id_tipo_documento' not in headers) and ('tipo_documento' not in headers):
            return Response({'error': 'Columna requerida faltante: fk_id_tipo_documento o tipo_documento'}, status=400)

        created, updated, errors = 0, 0, []
        header_idx = {h: i for i, h in enumerate(headers)}

        # Helper para resolver IDs por nombre en catálogos
        def resolve_by_name(Model, value):
            if value is None:
                return None
            
            # Si es número (o string numérico), úsalo como ID
            if str(value).isdigit():
                return int(value)
                
            s = str(value).strip()
            if not s:
                return None
                
            # Para Ciudad, buscar por nombre y retornar codigo_municipio
            if Model.__name__ == 'Ciudad':
                obj = Model.objects.filter(nombre__iexact=s).first()
                if obj:
                    return obj.codigo_municipio
                return None
                
            # Para otros modelos, buscar en campos estándar
            for field in ['descripcion', 'nombre', 'sigla', 'codigo']:
                if field in [f.name for f in Model._meta.fields]:
                    obj = Model.objects.filter(**{f"{field}__iexact": s}).first()
                    if obj:
                        return obj.pk
            return None

        for row in ws.iter_rows(min_row=2):
            try:
                data = {h: (row[header_idx[h]].value if header_idx.get(h) is not None else None) for h in headers}
                doc = str(data.get('numero_documento_estudiante')).strip()
                if not doc:
                    continue

                # Parse fecha_nacimiento si existe
                fecha_val = data.get('fecha_nacimiento(YYYY-MM-DD)') or data.get('fecha_nacimiento')
                if isinstance(fecha_val, datetime.date):
                    fecha_nacimiento = fecha_val
                elif isinstance(fecha_val, str) and fecha_val:
                    try:
                        fecha_nacimiento = datetime.datetime.strptime(fecha_val, '%Y-%m-%d').date()
                    except Exception:
                        fecha_nacimiento = None
                else:
                    fecha_nacimiento = None

                # Resolver FKs por nombre si vienen en lenguaje natural
                tipo_doc_id = normalize_fk(data.get('fk_id_tipo_documento') or data.get('tipo_documento'), TipoDocumento)
                genero_id = normalize_fk(data.get('fk_id_genero') or data.get('genero'), Genero)
                municipio_id = normalize_fk(data.get('fk_codigo_municipio') or data.get('municipio_id') or data.get('municipio'), Ciudad)
                tipo_sangre_id = normalize_fk(data.get('fk_id_tipo_sangre') or data.get('tipo_sangre'), TipoSangre)
                tipo_sisben_id = normalize_fk(data.get('fk_id_tipo_sisben') or data.get('sisben'), Sisben)
                estado_id = normalize_fk(data.get('fk_tipo_estado') or data.get('estado'), TipoEstado)

                defaults = {
                    'nombre1': data.get('nombre1') or '',
                    'nombre2': data.get('nombre2') or None,
                    'apellido1': data.get('apellido1') or '',
                    'apellido2': data.get('apellido2') or None,
                    'correo': data.get('correo') or None,
                    'direccion': data.get('direccion') or None,
                    'telefono': data.get('telefono') or None,
                    'religion': data.get('religion') or None,
                    'fecha_nacimiento': fecha_nacimiento,
                    'fk_id_tipo_documento_id': tipo_doc_id,
                    'fk_id_genero_id': genero_id,
                    'fk_codigo_municipio_id': municipio_id,
                    'fk_id_tipo_sangre_id': tipo_sangre_id,
                    'fk_id_tipo_sisben_id': tipo_sisben_id,
                    'fk_tipo_estado_id': estado_id,
                }

                obj, created_flag = Estudiantes.objects.update_or_create(
                    numero_documento_estudiante=doc,
                    defaults=defaults
                )
                created += 1 if created_flag else 0
                updated += 0 if created_flag else 1
            except Exception as e:
                errors.append({'row': row[0].row, 'error': str(e)})

        # Procesar hoja de Acudientes si existe
        acu_created, acu_updated, rel_created, acu_errors = 0, 0, 0, []
        if 'Acudientes' in wb.sheetnames:
            ws_acu = wb['Acudientes']
            headers_acu = [cell.value for cell in next(ws_acu.iter_rows(min_row=1, max_row=1))]
            required_acu = ['numero_documento_acudiente', 'nombre1', 'apellido1', 'numero_documento_estudiante']
            for req in required_acu:
                if req not in headers_acu:
                    acu_errors.append({'row': 1, 'error': f'Columna requerida faltante acudientes: {req}'})
            header_idx_acu = {h: i for i, h in enumerate(headers_acu)}

            for row in ws_acu.iter_rows(min_row=2):
                try:
                    data = {h: (row[header_idx_acu[h]].value if header_idx_acu.get(h) is not None else None) for h in headers_acu}
                    doc_acu = str(data.get('numero_documento_acudiente') or '').strip()
                    doc_est = str(data.get('numero_documento_estudiante') or '').strip()
                    if not doc_acu or not doc_est:
                        continue

                    # Resolver FKs acudiente por nombre si aplica
                    tipo_doc_acu_id = normalize_fk(data.get('fk_id_tipo_documento') or data.get('tipo_documento'), TipoDocumento)
                    municipio_acu_id = normalize_fk(data.get('fk_codigo_municipio') or data.get('municipio_id') or data.get('municipio'), Ciudad)
                    tipo_acudiente_id = normalize_fk(data.get('fk_id_tipo_acudiente') or data.get('tipo_acudiente'), TipoAcudiente)

                    defaults_acu = {
                        'fk_id_tipo_documento_id': tipo_doc_acu_id,
                        'nombre1': data.get('nombre1') or '',
                        'nombre2': data.get('nombre2') or None,
                        'apellido1': data.get('apellido1') or '',
                        'apellido2': data.get('apellido2') or None,
                        'telefono1': data.get('telefono1') or None,
                        'telefono2': data.get('telefono2') or None,
                        'direccion': data.get('direccion') or None,
                        'fk_codigo_municipio_id': municipio_acu_id,
                    }

                    acu_obj, acu_created_flag = Acudiente.objects.update_or_create(
                        numero_documento_acudiente=doc_acu,
                        defaults=defaults_acu
                    )
                    acu_created += 1 if acu_created_flag else 0
                    acu_updated += 0 if acu_created_flag else 1

                    # Crear/actualizar usuario core para el acudiente con credenciales basadas en documento
                    try:
                        # Si el usuario no existe, crearlo con contraseña igual al documento
                        user_qs = User.objects.filter(username=doc_acu)
                        if not user_qs.exists():
                            user = User.objects.create_user(
                                username=doc_acu,
                                password=doc_acu,
                                rol='padres',
                                first_name=str(data.get('nombre1') or ''),
                                last_name=(f"{data.get('apellido1') or ''} {data.get('apellido2') or ''}").strip(),
                                email=str(data.get('correo') or '')
                            )
                        else:
                            # Mantener sincronizados nombre, apellidos y correo; no cambiamos contraseña aquí
                            user_qs.update(
                                rol='padres',
                                first_name=str(data.get('nombre1') or ''),
                                last_name=(f"{data.get('apellido1') or ''} {data.get('apellido2') or ''}").strip(),
                                email=str(data.get('correo') or '')
                            )
                    except Exception as e:
                        # Registrar el error pero no detener el proceso de carga
                        acu_errors.append({'row': row[0].row, 'error': f'Error creando usuario core para acudiente {doc_acu}: {str(e)}'})

                    # Vincular acudiente con estudiante
                    # Normalizar documento del estudiante vinculado
                    try:
                        est_obj = None
                        if doc_est:
                            doc_est_str = str(doc_est).strip()
                            if doc_est_str:
                                try:
                                    est_obj = Estudiantes.objects.get(numero_documento_estudiante=doc_est_str)
                                except Estudiantes.DoesNotExist:
                                    # Intentar crear un estudiante placeholder mínimo si no existe (solo número y nombre genérico)
                                    est_obj = Estudiantes.objects.create(
                                        numero_documento_estudiante=doc_est_str,
                                        nombre1='N/A',
                                        apellido1='N/A'
                                    )
                        if est_obj is None:
                            acu_errors.append({'row': row[0].row, 'error': f'Estudiante {doc_est} no existe y no pudo crearse placeholder'})
                        else:
                            rel_defaults = {
                                'fk_id_tipo_acudiente_id': tipo_acudiente_id
                            }
                            rel_obj, rel_created_flag = EstudiantesAcudientes.objects.update_or_create(
                                fk_numero_documento_estudiante=est_obj,
                                fk_numero_documento_acudiente=acu_obj,
                                defaults=rel_defaults
                            )
                            rel_created += 1 if rel_created_flag else 0
                    except Exception as e:
                        acu_errors.append({'row': row[0].row, 'error': str(e)})
                except Exception as e:
                    acu_errors.append({'row': row[0].row, 'error': str(e)})

        # Log para depuración: imprimir errores procesados en consola del servidor
        try:
            if errors:
                print('DEBUG bulk-upload errors:', errors)
            if acu_errors:
                print('DEBUG bulk-upload acudientes errors:', acu_errors)
        except Exception as _e:
            print('DEBUG: no se pudo imprimir errores en consola:', _e)

        return Response({
            'created': created,
            'updated': updated,
            'errors': errors,
            'acudientes': {
                'created': acu_created,
                'updated': acu_updated,
                'relaciones_creadas': rel_created,
                'errors': acu_errors
            }
        })

class ProfesoresBulkUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if load_workbook is None:
            return Response({'error': 'openpyxl no está instalado. pip install openpyxl'}, status=500)
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Archivo "file" es requerido'}, status=400)

        try:
            wb = load_workbook(filename=file, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Error leyendo Excel: {str(e)}'}, status=400)

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        # Aceptar nombres naturales además de IDs para columnas de FK
        required = ['numero_documento_profesor', 'nombre1', 'apellido1', 'correo']
        for req in required:
            if req not in headers:
                return Response({'error': f'Columna requerida faltante: {req}'}, status=400)
        if ('fk_id_tipo_documento' not in headers) and ('tipo_documento' not in headers):
            return Response({'error': 'Columna requerida faltante: fk_id_tipo_documento o tipo_documento'}, status=400)

        created, updated, errors = 0, 0, []
        header_idx = {h: i for i, h in enumerate(headers)}

        for row in ws.iter_rows(min_row=2):
            try:
                data = {h: (row[header_idx[h]].value if header_idx.get(h) is not None else None) for h in headers}
                doc = str(data.get('numero_documento_profesor')).strip()
                if not doc:
                    continue
                # Normalizar y resolver FKs por nombre si vienen en lenguaje natural
                tipo_doc_id = normalize_fk(data.get('fk_id_tipo_documento') or data.get('tipo_documento'), TipoDocumento)
                municipio_id = normalize_fk(data.get('fk_codigo_municipio') or data.get('municipio'), Ciudad)
                estado_id = normalize_fk(data.get('fk_id_estado') or data.get('fk_tipo_estado') or data.get('estado'), TipoEstado)
                tipo_sangre_id = normalize_fk(data.get('fk_id_tipo_sangre') or data.get('tipo_sangre'), TipoSangre)

                defaults = {
                    'nombre1': data.get('nombre1') or '',
                    'nombre2': data.get('nombre2') or None,
                    'apellido1': data.get('apellido1') or '',
                    'apellido2': data.get('apellido2') or None,
                    'correo': data.get('correo') or None,
                    'direccion': data.get('direccion') or None,
                    'telefono1': data.get('telefono1') or None,
                    'telefono2': data.get('telefono2') or None,
                    'fk_id_tipo_documento_id': tipo_doc_id,
                    'fk_codigo_municipio_id': municipio_id,
                    'fk_id_estado_id': estado_id,
                    'fk_id_tipo_sangre_id': tipo_sangre_id,
                }
                obj, created_flag = Profesores.objects.update_or_create(
                    numero_documento_profesor=doc,
                    defaults=defaults
                )
                created += 1 if created_flag else 0
                updated += 0 if created_flag else 1
            except Exception as e:
                errors.append({'row': row[0].row, 'error': str(e)})

        return Response({'created': created, 'updated': updated, 'errors': errors})

# Vistas existentes
class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = RegisterSerializer


"""class LoginView(APIView):
    def post(self, request):
        username = request.data.get("username")
        password = request.data.get("password")
        
        user = authenticate(username=username, password=password)

        if user is not None:
            serializer = RegisterSerializer(user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response({"error": "Credenciales inválidas"}, status=status.HTTP_401_UNAUTHORIZED)"""


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer
